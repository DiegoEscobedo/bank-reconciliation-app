"""
bank_parser.py — Parser unificado de estados de cuenta bancarios.

Detecta automáticamente el formato del CSV y delega al sub-parser
correspondiente. Actualmente soportados:

    • BBVA     — header en fila 0, columnas DEPÓSITOS / RETIROS
    • BANORTE  — fila 0 tiene el número de cuenta, header en fila 1,
                 columnas Cargo / Abono

Agrega soporte para nuevos bancos implementando un sub-parser que herede
de _BaseBankParser y registrándolo en BankParser._PARSERS.
"""

import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from config.settings import TIENDA_ABBREV as _TIENDA_ABBREV
from src.utils.date_utils import looks_like_date, parse_date_spanish
from src.utils.logger import get_logger

logger = get_logger(__name__)


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().upper()
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


# ════════════════════════════════════════════════════════════
# BASE
# ════════════════════════════════════════════════════════════

class _BaseBankParser:
    """Interfaz común para todos los sub-parsers bancarios."""

    BANK_NAME: str = "UNKNOWN"

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        raise NotImplementedError

    @staticmethod
    def _read_file(file_path: str) -> pd.DataFrame:
        """
        Lee el archivo bancario (CSV o Excel) y retorna un DataFrame sin header
        donde todas las celdas son strings.
        """
        path = str(file_path).lower()

        if path.endswith(".xlsx") or path.endswith(".xls"):
            # Leer Excel; convertir todo a str para que el resto del pipeline
            # funcione igual que con CSV.
            df = pd.read_excel(file_path, header=None, dtype=str)
            # pd.read_excel con dtype=str convierte fechas a '2026-02-24 00:00:00';
            # strip para limpiar espacios residuales.
            return df.fillna("")

        for encoding in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                return pd.read_csv(
                    file_path,
                    header=None,
                    dtype=str,
                    encoding=encoding,
                    on_bad_lines="skip",
                )
            except UnicodeDecodeError:
                continue
        raise ValueError(f"No se pudo leer el archivo bancario: {file_path}")


# ════════════════════════════════════════════════════════════
# BBVA
# ════════════════════════════════════════════════════════════

class _BBVAParser(_BaseBankParser):
    """
    Formato Banorte (exportado como estructura BBVA-like):
        Fila 0 (header): CUENTA | FECHA DE OPERACIÓN | FECHA | REFERENCIA |
                         DESCRIPCIÓN | COD. TRANSAC | SUCURSAL |
                         DEPÓSITOS | RETIROS | SALDO | MOVIMIENTO |
                         DESCRIPCIÓN DETALLADA | CHEQUE
        Filas 1..n: datos
        
    Usado por: Banorte (cuentas como 3478, 6614, etc.)
    """

    BANK_NAME = "BANORTE"

    _COL_CUENTA    = "CUENTA"
    _COL_FECHA     = "FECHA DE OPERACIÓN"
    _COL_DESC      = "DESCRIPCIÓN"
    _COL_DESC_DET  = "DESCRIPCIÓN DETALLADA"
    _COL_COD_TRANS = "COD. TRANSAC"
    _COL_DEPOSITOS = "DEPÓSITOS"
    _COL_RETIROS   = "RETIROS"

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        # La primera fila es el header real
        header = [str(c).strip() for c in raw.iloc[0]]
        df = raw.iloc[1:].copy()
        df.columns = header
        df = df.reset_index(drop=True)

        account_id = self._extract_account(df)
        logger.info("[BBVA] Cuenta detectada: %s", account_id)

        # Normalizar nombres de columnas: remover acentos y espacios extra
        import unicodedata as _unicodedata
        
        def _normalize_col_name(s: str) -> str:
            """Quitar acentos y espacios extra de nombres de columna."""
            s = str(s).strip().upper()
            s = "".join(
                c for c in _unicodedata.normalize("NFD", s)
                if _unicodedata.category(c) != "Mn"
            )
            return s
        
        # Crear mapeo normalized → actual
        normalized_to_actual = {
            _normalize_col_name(col): col for col in df.columns
        }
        
        # Renombrar a nombres canónicos, siendo flexible con acentos/espacios
        rename_map = {}
        for old_name, new_name in [
            (self._COL_FECHA,     "raw_date"),
            (self._COL_DESC,      "description"),
            (self._COL_DESC_DET,  "description_detail"),
            (self._COL_COD_TRANS, "raw_cod_transac"),
            (self._COL_DEPOSITOS, "raw_deposit"),
            (self._COL_RETIROS,   "raw_withdrawal"),
        ]:
            norm_old = _normalize_col_name(old_name)
            if norm_old in normalized_to_actual:
                actual_col = normalized_to_actual[norm_old]
                rename_map[actual_col] = new_name
                logger.debug(f"[BBVA] Mapeado {old_name!r} → {actual_col!r} → {new_name!r}")
        
        df = df.rename(columns=rename_map)

        # Validar que tenemos las columnas mínimas necesarias
        if "raw_date" not in df.columns:
            raise ValueError(
                f"[BBVA] No se encontró la columna de fecha. "
                f"Columnas disponibles: {df.columns.tolist()}\n"
                f"Se esperaba 'FECHA DE OPERACIÓN', encontradas: "
                f"{[c for c in df.columns if 'FECHA' in c.upper()]}"
            )
        if "raw_deposit" not in df.columns and "raw_withdrawal" not in df.columns:
            raise ValueError(
                f"[BBVA] No se encontraron columnas de depósitos/retiros. "
                f"Columnas disponibles: {df.columns.tolist()}"
            )

        # Filtrar filas sin fecha válida
        df = df[df["raw_date"].apply(looks_like_date)].copy()

        df["account_id"] = account_id
        df["bank"] = self.BANK_NAME

        keep = ["account_id", "bank", "raw_date",
                "description", "description_detail",
            "raw_deposit", "raw_withdrawal", "raw_cod_transac"]
        return df[[c for c in keep if c in df.columns]].reset_index(drop=True)

    def _extract_account(self, df: pd.DataFrame) -> str:
        if self._COL_CUENTA in df.columns:
            val = str(df[self._COL_CUENTA].iloc[0]).strip()
            return val.lstrip("'")          # quitar apóstrofo inicial
        return "UNKNOWN"


# ════════════════════════════════════════════════════════════
# BANORTE
# ════════════════════════════════════════════════════════════

class _BanorteParser(_BaseBankParser):
    """
    Formato BBVA (estructura alternativa):
        Fila 0: Cuenta | <número de cuenta> | ...
        Fila 1 (header): Fecha Operación | Concepto | Referencia |
                         Referencia Ampliada | Cargo | Abono | Saldo
        Filas 2..n: datos
        
    Nota: Esta estructura no se está usando actualmente.
    """

    BANK_NAME = "BBVA"

    _COL_FECHA    = "Fecha Operación"
    _COL_CONCEPTO = "Concepto"
    _COL_REF_AMP  = "Referencia Ampliada"
    _COL_CARGO    = "Cargo"
    _COL_ABONO    = "Abono"

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        account_id = self._extract_account(raw)
        logger.info("[BANORTE] Cuenta detectada: %s", account_id)

        # Fila 1 es el header real
        header = [str(c).strip() for c in raw.iloc[1]]
        df = raw.iloc[2:].copy()
        df.columns = header
        df = df.reset_index(drop=True)

        # Normalizar nombres de columnas (como en BBVA)
        import unicodedata as _unicodedata
        
        def _normalize_col_name(s: str) -> str:
            s = str(s).strip().upper()
            s = "".join(
                c for c in _unicodedata.normalize("NFD", s)
                if _unicodedata.category(c) != "Mn"
            )
            return s
        
        normalized_to_actual = {
            _normalize_col_name(col): col for col in df.columns
        }
        
        rename_map = {}
        for old_name, new_name in [
            (self._COL_FECHA,    "raw_date"),
            (self._COL_CONCEPTO, "description"),
            (self._COL_REF_AMP,  "description_detail"),
            (self._COL_CARGO,    "raw_withdrawal"),
            (self._COL_ABONO,    "raw_deposit"),
        ]:
            norm_old = _normalize_col_name(old_name)
            if norm_old in normalized_to_actual:
                actual_col = normalized_to_actual[norm_old]
                rename_map[actual_col] = new_name
        
        df = df.rename(columns=rename_map)

        # Validar columnas mínimas
        if "raw_date" not in df.columns:
            raise ValueError(
                f"[BANORTE] No se encontró columna de fecha. "
                f"Disponibles: {df.columns.tolist()}"
            )
        if "raw_deposit" not in df.columns and "raw_withdrawal" not in df.columns:
            raise ValueError(
                f"[BANORTE] No se encontraron Cargo/Abono. "
                f"Disponibles: {df.columns.tolist()}"
            )

        # Filtrar filas sin fecha válida
        df = df[df["raw_date"].apply(looks_like_date)].copy()

        df["account_id"] = account_id
        df["bank"] = self.BANK_NAME

        keep = ["account_id", "bank", "raw_date",
                "description", "description_detail",
                "raw_deposit", "raw_withdrawal"]
        return df[[c for c in keep if c in df.columns]].reset_index(drop=True)

    @staticmethod
    def _extract_account(raw: pd.DataFrame) -> str:
        # Fila 0, columna 1 tiene el número de cuenta
        try:
            val = str(raw.iloc[0, 1]).strip()
            return val if val not in ("", "nan") else "UNKNOWN"
        except Exception:
            return "UNKNOWN"


# SCOTIABANK

class _ScotiabankParser(_BaseBankParser):
    """
    Formato Scotiabank (Excel sin fila de header):
        Col 0  → tipo (CHQ…)
        Col 1  → moneda (MXN)
        Col 3  → número de cuenta completo
        Col 4  → fecha (YYYY-MM-DD HH:MM:SS)
        Col 5  → referencia
        Col 6  → monto (siempre positivo)
        Col 7  → 'Cargo' (retiro) | 'Abono' (depósito)
        Col 9  → descripción
        Col 13 → descripción detallada
    """

    BANK_NAME = "SCOTIABANK"

    _IDX_ACCOUNT = 3
    _IDX_DATE    = 4
    _IDX_REF     = 5
    _IDX_AMOUNT  = 6
    _IDX_TYPE    = 7
    _IDX_DESC    = 9
    _IDX_DETAIL  = 13

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        # Filtrar filas con monto vacío (filas en blanco)
        df = raw[raw.iloc[:, self._IDX_AMOUNT].str.strip().ne("")].copy()
        df = df.reset_index(drop=True)

        if df.empty:
            logger.warning("[SCOTIABANK] No hay filas válidas con monto")
            return pd.DataFrame(columns=["account_id", "bank", "raw_date", "description", 
                                        "description_detail", "raw_deposit", "raw_withdrawal"])

        account_id = str(df.iloc[0, self._IDX_ACCOUNT]).strip() if len(df) else "UNKNOWN"
        logger.info("[SCOTIABANK] Cuenta detectada: %s", account_id)

        # Validar que las columnas necesarias existan
        num_cols = df.shape[1]
        logger.info("[SCOTIABANK] Archivo tiene %d columnas", num_cols)
        
        # Si no existe la columna de detalle, usar string vacío
        if num_cols > self._IDX_DETAIL:
            detail = df.iloc[:, self._IDX_DETAIL].astype(str).str.strip().replace("nan", "")
        else:
            logger.warning("[SCOTIABANK] Columna detalle (idx=%d) no existe (total=%d), usando vacío", 
                          self._IDX_DETAIL, num_cols)
            detail = pd.Series("", index=df.index)

        result = pd.DataFrame({
            "account_id":          df.iloc[:, self._IDX_ACCOUNT].astype(str).str.strip(),
            "bank":                self.BANK_NAME,
            "raw_date":            df.iloc[:, self._IDX_DATE].astype(str).str.strip(),
            "description":         df.iloc[:, self._IDX_DESC].astype(str).str.strip(),
            "description_detail":  detail,
            "raw_deposit":         df.apply(
                lambda r: r.iloc[self._IDX_AMOUNT] if str(r.iloc[self._IDX_TYPE]).strip() == "Abono" else "",
                axis=1,
            ),
            "raw_withdrawal":      df.apply(
                lambda r: r.iloc[self._IDX_AMOUNT] if str(r.iloc[self._IDX_TYPE]).strip() == "Cargo" else "",
                axis=1,
            ),
        })

        return result.reset_index(drop=True)


# ════════════════════════════════════════════════════════════
# NETPAY
# ════════════════════════════════════════════════════════════

class _NetPayParser(_BaseBankParser):
    """
    Formato NetPay (Excel): encabezados en una fila variable (busca
    dinámicamente hasta la fila 30). Columnas detectadas por nombre:
        Fecha de Movimiento / Fecha  → raw_date
        Clave de Rastreo / Referencia → description
        Cuenta Destino / Cuenta Depósito → account_id
        Monto de Trx / Monto TRX / Monto Depósito → raw_deposit
    Todos los movimientos son DEPÓSITOS al banco destino.
    """

    BANK_NAME = "NETPAY"

    # Palabras clave para mapear columnas (orden: preferida → fallback)
    # Usamos FECHA TRX porque el JDE registra la fecha de la transacción,
    # no la fecha de depósito.
    _COL_DATE    = ("FECHA TRX", "FECHA DE MOVIMIENTO", "FECHA MOVIMIENTO", "FECHA DE DEPÓSITO", "FECHA DE DEPOSITO", "FECHA")
    _COL_ACCOUNT = ("CUENTA DESTINO", "CUENTA DEPÓSITO", "CUENTA DEPOSITO", "CUENTA")
    _COL_DESC    = ("SUCURSAL", "DESCRIPCIÓN", "DESCRIPCION", "CLAVE RASTREO", "CLAVE DE RASTREO", "REFERENCIA")
    _COL_AMOUNT  = ("MONTO DE TRX", "MONTO DE TRANSACCIÓN", "MONTO TRX", "MONTO TRANSACCION", "MONTO DEPÓSITO", "MONTO DEPOSITO", "MONTO")

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        # ── 1. Detectar fila de encabezado ────────────────────────────
        header_row_idx = None
        col_map: dict[str, int] = {}

        for _ri in range(min(30, len(raw))):
            row_vals = raw.iloc[_ri].fillna("").astype(str)
            row_upper = [v.strip().upper() for v in row_vals]
            # Buscar al menos fecha + monto en esta fila
            has_fecha = any("FECHA" in v for v in row_upper)
            has_monto = any("MONTO" in v for v in row_upper)
            if has_fecha and has_monto:
                header_row_idx = _ri
                for _ci, _v in enumerate(row_upper):
                    col_map[_v] = _ci
                break

        if header_row_idx is None:
            logger.warning("[NETPAY] No se encontró fila de encabezado, usando defaults.")
            header_row_idx = 9
            # Fallback a índices fijos legacy
            col_map = {}

        def _find_col(keywords, default=None):
            for kw in keywords:
                if kw in col_map:
                    return col_map[kw]
            # Búsqueda parcial
            for kw in keywords:
                for k, v in col_map.items():
                    if kw in k:
                        return v
            return default

        idx_date    = _find_col(self._COL_DATE,    default=2)
        idx_account = _find_col(self._COL_ACCOUNT, default=4)
        idx_desc    = _find_col(self._COL_DESC,    default=5)
        idx_amount  = _find_col(self._COL_AMOUNT,  default=6)

        logger.info(
            "[NETPAY] Header fila=%d | fecha=%d cuenta=%d desc=%d monto=%d",
            header_row_idx, idx_date, idx_account, idx_desc, idx_amount,
        )

        # ── 2. Extraer filas de datos ─────────────────────────────────
        data = raw.iloc[header_row_idx + 1:].copy().reset_index(drop=True)
        data = data.fillna("").astype(str)

        # Filtrar filas vacías o de totales
        data = data[
            data.iloc[:, idx_date].str.strip().ne("") &
            data.iloc[:, idx_date].str.strip().str.upper().ne("TOTAL") &
            data.iloc[:, idx_amount].str.strip().ne("")
        ].copy().reset_index(drop=True)

        if data.empty:
            return pd.DataFrame(columns=["account_id", "bank", "raw_date",
                                         "description", "description_detail",
                                         "raw_deposit", "raw_withdrawal"])

        # account_id desde primera fila válida
        account_id = str(data.iloc[0, idx_account]).strip()
        logger.info("[NETPAY] Cuenta destino: %s", account_id)

        result = pd.DataFrame({
            "account_id":         account_id,
            "bank":               self.BANK_NAME,
            "raw_date":           data.iloc[:, idx_date].str.strip(),
            "description":        data.iloc[:, idx_desc].str.strip(),
            "description_detail": "",
            "raw_deposit":        data.iloc[:, idx_amount].str.strip(),
            "raw_withdrawal":     "",
        })
        result = result.reset_index(drop=True)
        # Mapear Sucursal → tienda (abreviatura JDE) usando TIENDA_ABBREV
        result["tienda"] = (
            data.iloc[:, idx_desc].str.strip().str.upper()
            .map(_TIENDA_ABBREV)
            .fillna("")
        )
        # NetPay no trae tipo de pago en su Excel origen; por negocio se trata como TPV.
        result["tipo_banco"] = "TPV"
        # Diagnóstico: mostrar primeras filas extraídas
        for _ri in range(min(3, len(result))):
            logger.info(
                "[NETPAY] Fila %d: raw_date=%r  raw_deposit=%r  desc=%r",
                _ri, result.at[_ri, "raw_date"],
                result.at[_ri, "raw_deposit"],
                result.at[_ri, "description"],
            )
        return result


# ════════════════════════════════════════════════════════════
# MERCADO PAGO
# ════════════════════════════════════════════════════════════

class _MercadoPagoParser(_BaseBankParser):
    """
    Formato Mercado Pago (Excel):
        Fila 3  → header con 38 columnas
        Fila 4+ → datos
    
    NUEVA LÓGICA: Lee colores de la celda en columna "Total a recibir" (col8).
    Conserva cualquier color (incluyendo blanco/gris) y genera una clave
    estable para agrupar correctamente por color.
    
    Columnas usadas (buscadas por nombre en la fila de cabecera):
        'Número de operación'  → folio
        'Fecha de la compra'   → fecha  ('18 feb 19:14 hs')
        'Cobro'                → monto bruto cobrado al comprador
        'Total a recibir'      → **COLUMNA CON COLORES** (para agrupar)
        Sucursal/descripción   → última columna no-vacía o col34
    
    Todos son DEPÓSITOS. account_id = "7133" (Scotiabank destino).
    """

    BANK_NAME        = "MERCADOPAGO"
    _DESTINATION_ACT = "7133"   # Scotiabank donde deposita MP

    _HEADER_ROW  = 3

    # Nombres de columna a buscar (en orden de prioridad, case-insensitive)
    _COL_FOLIO    = ("NÚMERO DE OPERACIÓN", "NUMERO DE OPERACION", "OPERACION", "FOLIO")
    _COL_DATE     = ("FECHA DE LA COMPRA",  "FECHA DE OPERACIÓN",  "FECHA DE OPERACION", "FECHA")
    _COL_AMOUNT   = ("COBRO",)                          # monto bruto
    _COL_TOTAL_RECIBIR = ("TOTALARECIBIR", "TOTAL A RECIBIR", "TOTAL ARECIBIR")  # para colores
    _COL_STATUS   = ("ESTADO", "STATUS")                # filtro: solo "Aprobado"
    _COL_SUCURSAL = ("SUCURSAL", "DESCRIPCIÓN", "DESCRIPCION")

    # Índices de respaldo
    _FALLBACK_IDX_FOLIO    = 0
    _FALLBACK_IDX_DATE     = 1
    _FALLBACK_IDX_AMOUNT   = 4
    _FALLBACK_IDX_TOTAL_RECIBIR = 8  # col8 = Total a recibir
    _FALLBACK_IDX_STATUS   = 6       # col6 = Estado (por defecto)
    _FALLBACK_IDX_SUCURSAL = 34

    def _resolve_col(self, headers: list[str], names: tuple, fallback: int) -> int:
        """Devuelve el índice de la primera columna cuyo nombre (normalizado) coincida."""
        normalized = [h.strip().upper() for h in headers]
        for name in names:
            try:
                return normalized.index(name.upper())
            except ValueError:
                continue
        return fallback

    def _extract_color_key(self, fill: PatternFill) -> str:
        """
        Devuelve una clave de color estable sin ignorar blanco/gris.
        Prioriza RGB y, si no existe, usa type/indexed/theme/auto.
        """
        if not fill or not fill.fgColor:
            return "NO_FILL"

        fg = fill.fgColor
        try:
            color_type = str(getattr(fg, "type", "") or "").strip().lower()
        except (AttributeError, TypeError, ValueError):
            color_type = ""

        rgb = getattr(fg, "rgb", None)
        if rgb:
            return f"RGB:{str(rgb).strip().upper()}"

        indexed = getattr(fg, "indexed", None)
        if indexed is not None:
            return f"INDEXED:{indexed}"

        theme = getattr(fg, "theme", None)
        tint = getattr(fg, "tint", None)
        if theme is not None:
            tint_key = "" if tint is None else f":TINT:{tint}"
            return f"THEME:{theme}{tint_key}"

        auto = getattr(fg, "auto", None)
        if auto is not None:
            return f"AUTO:{auto}"

        value = getattr(fg, "value", None)
        if value:
            return f"{color_type.upper() or 'COLOR'}:{str(value).strip().upper()}"

        return "UNKNOWN_COLOR"

    def _normalize_amount(self, amount_str: str) -> str:
        """Normaliza monto: remueve $, convierte coma a punto, etc."""
        # Remover espacios y símbolo $
        normalized = amount_str.strip().replace("$", "").strip()
        # Cambiar coma por punto (formato español → formato float)
        normalized = normalized.replace(",", ".")
        # Remover espacios internos
        normalized = normalized.replace(" ", "")
        return normalized

    def _get_cell_color_hex(self, cell) -> str:
        """Extrae una clave de color para agrupar filas de Mercado Pago."""
        try:
            fill = cell.fill
            return self._extract_color_key(fill)
        except (AttributeError, TypeError):
            pass
        return "UNKNOWN_COLOR"

    def parse(self, file_path: str) -> pd.DataFrame:
        """Parsea Mercado Pago desde Excel, extrayendo colores de las celdas."""
        logger.info("[MERCADOPAGO] Parseando: %s", file_path)
        
        try:
            # Leer con pandas para obtener datos
            raw = pd.read_excel(file_path, header=None, dtype=str).fillna("")
            
            # Leer con openpyxl para obtener colores
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Encontrar columna de "Total a recibir" para leer colores
            header_row = raw.iloc[self._HEADER_ROW].astype(str).tolist()
            idx_total_recibir = self._resolve_col(
                header_row,
                self._COL_TOTAL_RECIBIR,
                self._FALLBACK_IDX_TOTAL_RECIBIR,
            )
            # En Excel, las columnas son 1-indexed
            col_letter_total = chr(65 + idx_total_recibir)
            
            # Resolver otros índices
            idx_folio    = self._resolve_col(header_row, self._COL_FOLIO,    self._FALLBACK_IDX_FOLIO)
            idx_date     = self._resolve_col(header_row, self._COL_DATE,     self._FALLBACK_IDX_DATE)
            idx_amount   = self._resolve_col(header_row, self._COL_AMOUNT,   self._FALLBACK_IDX_AMOUNT)
            idx_status   = self._resolve_col(header_row, self._COL_STATUS,   self._FALLBACK_IDX_STATUS)
            idx_sucursal = self._resolve_col(header_row, self._COL_SUCURSAL, self._FALLBACK_IDX_SUCURSAL)

            logger.info(
                "[MERCADOPAGO] Columnas resueltas → folio=%d fecha=%d cobro=%d total_recibir=%d status=%d sucursal=%d",
                idx_folio, idx_date, idx_amount, idx_total_recibir, idx_status, idx_sucursal,
            )
            
            # Debug: mostrar las columnas detectadas
            logger.debug("[MERCADOPAGO] Header completo: %s", header_row[:15])

            # Iterar filas de datos y extraer colores
            data_rows = []
            skipped_rows = []  # Para logging
            for excel_row_idx in range(self._HEADER_ROW + 2, ws.max_row + 1):  # +2 porque Excel es 1-indexed y saltamos header+1
                cell_color = self._get_cell_color_hex(
                    ws[f"{col_letter_total}{excel_row_idx}"]
                )
                # La selección final de filas válidas por color se hace en main.py
                # (grupos cuyo total aparece en Scotiabank). Aquí no descartamos por color.
                cell_color = cell_color or "UNKNOWN_COLOR"
                
                # Obtener datos de la fila desde pandas
                # Conversión: Excel usa 1-indexed, pandas usa 0-indexed
                pandas_row_idx = excel_row_idx - 1
                if pandas_row_idx >= len(raw):
                    break
                
                row_data = raw.iloc[pandas_row_idx]
                cobro_str = str(row_data.iloc[idx_amount]).strip()
                total_recibir_str = str(row_data.iloc[idx_total_recibir]).strip()
                cobro_normalized = self._normalize_amount(cobro_str)
                total_recibir_normalized = self._normalize_amount(total_recibir_str)
                # Para match contra JDE se usa COBRO (requerimiento funcional).
                amount_normalized = cobro_normalized
                status_text = str(row_data.iloc[idx_status]).strip().lower()
                folio_text = str(row_data.iloc[idx_folio]).strip()
                
                # Filtrar filas sin monto válido
                if amount_normalized in ("", "nan", "0", "0.0"):
                    skipped_rows.append((excel_row_idx, f"EMPTY_COBRO:cobro={cobro_str}|total={total_recibir_str}"))
                    logger.debug("[MERCADOPAGO] Fila %d ignorada: monto vacío o 0", excel_row_idx)
                    continue
                
                # Validar que el monto sea convertible a float
                try:
                    float(amount_normalized)
                except ValueError:
                    skipped_rows.append((excel_row_idx, f"INVALID_AMOUNT:{amount_normalized}"))
                    logger.debug(
                        "[MERCADOPAGO] Fila %d ignorada: COBRO inválido cobro='%s' total='%s' -> '%s'",
                        excel_row_idx, cobro_str, total_recibir_str, amount_normalized,
                    )
                    continue
                
                # Filtrar filas sin estado aprobado (Aprobado/Aprovado)
                is_approved_status = ("aprobado" in status_text) or ("aprovado" in status_text)
                if not is_approved_status:
                    skipped_rows.append((excel_row_idx, f"STATUS:{status_text}"))
                    logger.debug("[MERCADOPAGO] Fila %d ignorada: Estado='%s' (no Aprobado)", excel_row_idx, status_text)
                    continue
                
                logger.debug("[MERCADOPAGO] Fila %d PROCESADA: folio=%s monto=%s status=%s color=%s",
                           excel_row_idx, folio_text, amount_normalized, status_text, cell_color)
                
                data_rows.append({
                    "excel_row_idx": excel_row_idx,
                    "cell_color": cell_color,
                    "folio": folio_text,
                    "raw_date": str(row_data.iloc[idx_date]).strip(),
                    "raw_amount": amount_normalized,  # COBRO normalizado
                    "raw_cobro": cobro_normalized,
                    "raw_total_recibir": total_recibir_normalized,
                    "raw_status": status_text,
                    "raw_sucursal": str(row_data.iloc[idx_sucursal]).strip().replace("nan", ""),
                })
            
            logger.info("[MERCADOPAGO] Resumen de parseo: %d rows procesadas, %d rows descartadas", 
                       len(data_rows), len(skipped_rows))
            if skipped_rows:
                logger.debug("[MERCADOPAGO] Filas descartadas: %s", skipped_rows[:10])
            
            if not data_rows:
                logger.warning("[MERCADOPAGO] No se encontraron filas válidas (monto/estado)")
                return pd.DataFrame(columns=["account_id", "bank", "raw_date",
                                             "description", "description_detail",
                                             "raw_deposit", "raw_withdrawal", "cell_color"])
            
            # Convertir a DataFrame
            data = pd.DataFrame(data_rows)
            
            logger.info("[MERCADOPAGO] %d filas con color encontradas. Colores únicos: %s",
                        len(data), data["cell_color"].nunique())
            logger.info("[MERCADOPAGO] Colores detectados: %s", sorted(data["cell_color"].unique()))
            
            # Convertir fechas
            dates_fmt = data["raw_date"].apply(
                lambda v: parse_date_spanish(v).strftime("%d/%m/%Y")
                if parse_date_spanish(v) is not pd.NaT else v
            )
            
            result = pd.DataFrame({
                "account_id":         self._DESTINATION_ACT,
                "bank":               self.BANK_NAME,
                "raw_date":           dates_fmt,
                "description":        "MERCADO PAGO | " + data["raw_sucursal"],
                "description_detail": data["folio"],
                "raw_deposit":        data["raw_amount"],
                "raw_withdrawal":     "",
                "cell_color":         data["cell_color"],  # color de celda para grouping
                "raw_cobro":          data["raw_cobro"],
                "raw_total_recibir":  data["raw_total_recibir"],
                "raw_status":         data["raw_status"],   # Estado (Aprobado, etc)
                "_excel_row_idx":     data["excel_row_idx"],  # para write-back
                "_mp_row_type":       "COBRO",
            })
            
            # Mapear tienda
            result["tienda"] = (
                data["raw_sucursal"].str.upper()
                .map(_TIENDA_ABBREV)
                .fillna("")
            )
            # Mercado Pago no trae tipo de pago explícito; por negocio se trata como TPV.
            result["tipo_banco"] = "TPV"

            # Generar movimientos de COMISION como retiros: COBRO - TOTAL A RECIBIR.
            # Esto permite que varias comisiones MP se agrupen contra un solo JDE de comision.
            cobro_num = pd.to_numeric(result["raw_cobro"], errors="coerce").fillna(0.0)
            total_recibir_num = pd.to_numeric(result["raw_total_recibir"], errors="coerce").fillna(0.0)
            commission_amount = (cobro_num - total_recibir_num).round(2)
            commission_mask = commission_amount > 0

            commission_count = int(commission_mask.sum())
            commission_total = float(commission_amount[commission_mask].sum()) if commission_count else 0.0

            if commission_count > 0:
                commission_df = result.loc[commission_mask].copy()
                commission_df["description"] = (
                    "MERCADO PAGO | COMISION | "
                    + commission_df["description"].fillna("").astype(str)
                )
                commission_df["raw_deposit"] = ""
                commission_df["raw_withdrawal"] = commission_amount[commission_mask].map(lambda v: f"{v:.2f}").values
                commission_df["_mp_row_type"] = "COMISION"
                # Estas filas no deben influir en el filtro por total de color.
                commission_df["raw_total_recibir"] = ""

                result = pd.concat([result, commission_df], ignore_index=True)

            logger.info(
                "[MERCADOPAGO] Comisiones generadas: %d filas, total=%.2f",
                commission_count,
                commission_total,
            )
            
            logger.info("[MERCADOPAGO] Movimientos procesados: %d", len(result))
            logger.info("[MERCADOPAGO] Tiendas detectadas: %s", result["tienda"].value_counts().to_dict())
            
            return result.reset_index(drop=True)
            
        except Exception as e:
            logger.error("[MERCADOPAGO] Error parseando: %s", str(e))
            return pd.DataFrame(columns=["account_id", "bank", "raw_date",
                                         "description", "description_detail",
                                         "raw_deposit", "raw_withdrawal", "cell_color"])

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        """Este método ya no se usa; parse() es el punto de entrada."""
        logger.warning("[MERCADOPAGO] parse_raw() fue llamado; use parse(file_path) en su lugar")
        return pd.DataFrame()


# ════════════════════════════════════════════════════════════
# REPORTE (banco enriquecido con tienda+tipo por el usuario)
# ════════════════════════════════════════════════════════════

class _ReporteParser(_BaseBankParser):
    """
    Formato REPORTE.xlsx — Hoja2:
        Fila 0 (header): CUENTA | FECHA DE OPERACIÓN | FECHA | REFERENCIA |
                         DESCRIPCIÓN | COD. TRANSAC | SUCURSAL |
                         DEPÓSITOS | RETIROS | SALDO | MOVIMIENTO |
                         DESCRIPCIÓN DETALLADA | CHEQUE | TPV
        Filas 1..n: datos

    La columna 13 ("TPV") contiene la clasificación que el usuario agregó
    con fórmulas Excel: p. ej. "OUG TPV", "FAB TR", "FRE 03".
    Esta se parsea para extraer tienda y tipo_banco.
    """

    BANK_NAME = "REPORTE"

    _TIPOS_BANCO = {"TPV", "TR", "01", "03", "04", "28"}

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        # raw = Hoja2, fila 0 = cabeceras
        header = [str(c).strip() for c in raw.iloc[0]]
        df = raw.iloc[1:].copy()
        # Asignar nombres de columna; si hay más columnas que cabeceras, usar índice
        df.columns = [
            header[i] if i < len(header) else f"_col{i}"
            for i in range(len(raw.columns))
        ]
        df = df.reset_index(drop=True)

        # Cuenta (col 0 = CUENTA)
        raw_cuenta = df.get("CUENTA", df.iloc[:, 0]).astype(str)
        df["account_id"] = raw_cuenta.str.strip().str.lstrip("'")
        account_id = df["account_id"].mode()[0] if not df.empty else "UNKNOWN"
        logger.info("[REPORTE] Cuenta detectada: %s", account_id)

        # Fecha de liquidación (col 2 = FECHA)
        fecha_col = header[2] if len(header) > 2 else "FECHA"
        df["raw_date"] = df.get(fecha_col, df.iloc[:, 2]).astype(str)

        # Montos: col 7 = DEPÓSITOS, col 8 = RETIROS
        col_dep = header[7] if len(header) > 7 else "DEPÓSITOS"
        col_ret = header[8] if len(header) > 8 else "RETIROS"
        df["raw_deposit"]    = df.get(col_dep, df.iloc[:, 7]).astype(str)
        df["raw_withdrawal"] = df.get(col_ret, df.iloc[:, 8]).astype(str)

        # Descripción: col 4 y col 11
        col_desc = header[4] if len(header) > 4 else "DESCRIPCIÓN"
        col_det  = header[11] if len(header) > 11 else "DESCRIPCIÓN DETALLADA"
        df["description"]        = df.get(col_desc, df.iloc[:, 4]).astype(str)
        df["description_detail"] = (
            df.get(col_det, df.iloc[:, 11]).astype(str)
            if len(df.columns) > 11 else ""
        )

        # COD. TRANSAC: col 5
        col_cod = header[5] if len(header) > 5 else "COD. TRANSAC"
        df["raw_cod_transac"] = df.get(col_cod, df.iloc[:, 5]).astype(str)

        # Clasificación tienda+tipo (col 13 = TPV)
        col_tpv = header[13] if len(header) > 13 else "TPV"
        clasificacion = df.get(col_tpv, df.iloc[:, 13]).astype(str)
        parsed_tpv   = clasificacion.apply(self._parse_clasificacion)
        df["tienda"]     = parsed_tpv.apply(lambda x: x[0])
        df["tipo_banco"] = parsed_tpv.apply(lambda x: x[1])

        # Filtrar filas sin fecha válida
        df = df[df["raw_date"].apply(looks_like_date)].copy()
        df["account_id"] = account_id
        df["bank"]       = self.BANK_NAME

        keep = [
            "account_id", "bank", "raw_date",
            "description", "description_detail",
            "raw_deposit", "raw_withdrawal",
            "tienda", "tipo_banco", "raw_cod_transac",
        ]
        return df[[c for c in keep if c in df.columns]].reset_index(drop=True)

    def _parse_clasificacion(self, value: str):
        """Parsea "OUG TPV" → ("OUG", "TPV"),  "FRE 03" → ("FRE", "03")."""
        value = str(value).strip()
        if not value or value in ("", "nan", "-"):
            return (None, None)
        parts = value.split()
        if len(parts) >= 2:
            tipo = parts[-1].upper()
            if tipo in self._TIPOS_BANCO:
                tienda = " ".join(parts[:-1]).upper()
                return (tienda, tipo)
        return (None, None)


# ════════════════════════════════════════════════════════════
# REPORTE CAJA  (resumen diario por tienda: fecha|banco|monto|tienda|tipo)
# ════════════════════════════════════════════════════════════


class _ReporteCajaParser(_BaseBankParser):
    """
    Formato REPORTE CAJA.xlsx — Hoja1:
        Filas 0..N: título / vacías (cantidad variable)
        Fila N+1 (header): FECHA | Banco | Monto | TIENDA | OBSERVACION(ES)
        Filas N+2..n: datos

    El parser detecta automáticamente la fila de cabeceras buscando
    la palabra "FECHA" en la columna 0 o "TIENDA" en cualquier columna.

    Columnas:
        Col 0  → FECHA (fecha del depósito)
        Col 1  → Banco  (p. ej. "BANORTE 6614", "BBVA 4640")
        Col 2  → Monto  (depósito)
        Col 3  → TIENDA (nombre completo, p. ej. "OUTLET JALPA")
        Col 4  → OBSERVACION / OBSERVACIONES ("TR", "TPV" o vacío=efectivo)
    """

    BANK_NAME = "REPORTE_CAJA"

    # Palabras clave que identifican la fila de cabeceras
    _HEADER_KEYWORDS = {"FECHA", "TIENDA", "MONTO", "BANCO"}

    @staticmethod
    def _extract_tipo_banco_from_text(*parts) -> str | None:
        """
        Extrae tipo_banco desde texto libre.
        Prioriza TPV/TR explícitos y también reconoce "TRANSFERENCIA" como TR.
        """
        text = " ".join(str(p) for p in parts if p is not None).upper()
        if re.search(r"(?<![A-Z0-9])TPV(?![A-Z0-9])", text):
            return "TPV"
        if re.search(r"(?<![A-Z0-9])TR(?![A-Z0-9])", text) or "TRANSFERENCIA" in text:
            return "TR"
        return None

    def _find_header_row(self, raw: pd.DataFrame) -> int:
        """Devuelve el índice (0-based) de la fila de cabeceras."""
        for i, row in raw.iterrows():
            row_vals = {str(v).strip().upper() for v in row.values if str(v).strip()}
            if len(row_vals & self._HEADER_KEYWORDS) >= 2:
                return i
        return 0  # fallback: primera fila

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        header_idx = self._find_header_row(raw)
        # La fila de cabeceras se usa como nombres de columna
        header_row  = raw.iloc[header_idx]
        df = raw.iloc[header_idx + 1:].copy().reset_index(drop=True)
        df.columns = [str(c).strip() for c in header_row.values]

        # Resolver nombres de columna con variantes posibles
        def _col(candidates):
            """Devuelve el nombre de columna que coincida (case-insensitive)."""
            for c in df.columns:
                if c.strip().upper() in {x.upper() for x in candidates}:
                    return c
            return None

        col_fecha  = _col(["FECHA"])
        col_banco  = _col(["BANCO", "BANK"])
        col_monto  = _col(["MONTO", "IMPORTE", "AMOUNT"])
        col_tienda = _col(["TIENDA", "SUCURSAL", "STORE"])
        col_obs    = _col(["OBSERVACION", "OBSERVACIONES", "OBS", "TIPO", "TIPO PAGO"])

        rows = []
        for _, row in df.iterrows():
            fecha      = str(row[col_fecha]).strip()  if col_fecha  else str(row.iloc[0]).strip()
            banco      = str(row[col_banco]).strip()  if col_banco  else str(row.iloc[1]).strip()
            monto      = str(row[col_monto]).strip()  if col_monto  else str(row.iloc[2]).strip()
            tienda_raw = str(row[col_tienda]).strip() if col_tienda else (str(row.iloc[3]).strip() if len(row) > 3 else "")
            obs        = str(row[col_obs]).strip()    if col_obs    else (str(row.iloc[4]).strip() if len(row) > 4 else "")

            # Saltar filas vacías o sin fecha/monto
            if not looks_like_date(fecha) or monto in ("", "nan", "0"):
                continue

            # Extraer account_id del campo Banco ("BANORTE 6614" → "6614")
            account_id = self._extract_account_from_banco(banco)

            # Mapear tienda a abreviatura JDE
            tienda_abbrev = _TIENDA_ABBREV.get(tienda_raw.upper(), "NO ENCONTRADO")

            # Tipo banco: detectar TR/TPV desde observacion/descripcion incluso sin tienda.
            obs_clean = obs.strip().upper()
            tipo_detectado = self._extract_tipo_banco_from_text(obs_clean, tienda_raw, banco)
            if not obs_clean:
                # Regla de negocio: observación vacía se considera efectivo.
                tipo_banco = "01"
            else:
                # Si hay texto en observación y no cae en TR/TPV, conservar texto crudo.
                tipo_banco = tipo_detectado or obs_clean

            rows.append({
                "account_id":        account_id,
                "bank":              self.BANK_NAME,
                "raw_date":          fecha,
                "description":       f"{tienda_raw} | {obs_clean}" if obs_clean else tienda_raw,
                "description_detail": banco,
                "raw_deposit":       monto,
                "raw_withdrawal":    "",
                "tienda":            tienda_abbrev,
                "tipo_banco":        tipo_banco,
            })

        result = pd.DataFrame(rows)
        if result.empty:
            return result

        # Stats sobre tiendas mapeadas vs no encontradas
        tiendas_mapeadas = (result["tienda"] != "NO ENCONTRADO").sum()
        tiendas_no_encontradas = (result["tienda"] == "NO ENCONTRADO").sum()

        # Agrupar multiples cuentas por si el reporte mezcla (6614, 4640…)
        accounts = result["account_id"].unique()
        logger.info("[REPORTE_CAJA] Cuentas detectadas: %s", accounts.tolist())
        logger.info("[REPORTE_CAJA] Tiendas mapeadas: %d, No encontradas: %d", tiendas_mapeadas, tiendas_no_encontradas)
        
        return result.reset_index(drop=True)

    @staticmethod
    def _extract_account_from_banco(banco: str) -> str:
        """"BANORTE 6614" → "6614",  "BBVA 4640" → "4640"."""
        parts = banco.split()
        if parts:
            return parts[-1]  # último token es el número de cuenta
        return "UNKNOWN"


# ════════════════════════════════════════════════════════════
# HSBC
# ════════════════════════════════════════════════════════════

class _HSBCParser(_BaseBankParser):
    """
    Formato HSBC:
        Fila 0 (header):
            Col 1: Nombre de cuenta
            Col 2: Número de cuenta
            Col 3: Nombre del banco (HSBC Mexico)
            ...
            Col 19: Descripción
            Col 22: Fecha valor
            Col 23: Importe de crédito
            Col 24: Importe del débito
        Filas 1..n: datos
    
    Estructura simple: headers en row 0, datos a partir de row 1.
    """

    BANK_NAME = "HSBC"

    _COL_CUENTA = 1      # 0-indexed: Col 2
    _COL_DESC = 18       # 0-indexed: Col 19
    _COL_FECHA = 21      # 0-indexed: Col 22
    _COL_CREDITO = 22    # 0-indexed: Col 23 (depósito)
    _COL_DEBITO = 23     # 0-indexed: Col 24 (retiro)

    _HEADER_ACCOUNT = (
        "NUMERO DE CUENTA",
        "NUMERO CUENTA",
        "CUENTA",
    )
    _HEADER_DESC = (
        "DESCRIPCION",
    )
    _HEADER_FECHA = (
        "FECHA VALOR",
        "FECHA DEL APUNTE",
        "FECHA",
    )
    _HEADER_CREDITO = (
        "IMPORTE DE CREDITO",
        "CREDITO",
    )
    _HEADER_DEBITO = (
        "IMPORTE DEL DEBITO",
        "IMPORTE DE DEBITO",
        "DEBITO",
    )

    @staticmethod
    def _resolve_col(headers: list[str], keywords: tuple[str, ...], fallback: int) -> str:
        normalized_to_actual = {
            _normalize_text(col): col
            for col in headers
        }
        for kw in keywords:
            kw_norm = _normalize_text(kw)
            if kw_norm in normalized_to_actual:
                return normalized_to_actual[kw_norm]
        for kw in keywords:
            kw_norm = _normalize_text(kw)
            for norm, actual in normalized_to_actual.items():
                if kw_norm in norm:
                    return actual
        if 0 <= fallback < len(headers):
            return headers[fallback]
        return ""

    def parse_raw(self, raw: pd.DataFrame) -> pd.DataFrame:
        # Las columnas ya están numeradas (0-indexed) desde el read_file
        # Fila 0 es header, filas 1+ son datos
        if len(raw) < 2:
            logger.warning("[HSBC] Archivo vacío o sin datos")
            return pd.DataFrame(columns=[
                "account_id", "bank", "raw_date",
                "description", "description_detail",
                "raw_deposit", "raw_withdrawal"
            ])

        # Extraer headers
        headers = [str(c).strip() for c in raw.iloc[0]]
        
        # Datos desde la fila 1
        df = raw.iloc[1:].copy()
        df.columns = headers
        df = df.reset_index(drop=True)

        col_cuenta_name = self._resolve_col(headers, self._HEADER_ACCOUNT, self._COL_CUENTA)
        col_desc_name = self._resolve_col(headers, self._HEADER_DESC, self._COL_DESC)
        col_fecha_name = self._resolve_col(headers, self._HEADER_FECHA, self._COL_FECHA)
        col_credito_name = self._resolve_col(headers, self._HEADER_CREDITO, self._COL_CREDITO)
        col_debito_name = self._resolve_col(headers, self._HEADER_DEBITO, self._COL_DEBITO)

        # Extraer cuenta desde la primera fila válida de la columna de cuenta.
        account_id = "UNKNOWN"
        if col_cuenta_name and col_cuenta_name in df.columns:
            cuentas = (
                df[col_cuenta_name]
                .fillna("")
                .astype(str)
                .str.strip()
            )
            cuentas = cuentas[
                ~cuentas.str.lower().isin(["", "nan", "none"])
            ]
            if not cuentas.empty:
                account_id = str(cuentas.iloc[0]).strip()

        logger.info("[HSBC] Cuenta detectada: %s", account_id)

        # Validar que tenemos columnas mínimas
        if not col_fecha_name or col_fecha_name not in df.columns:
            available = df.columns.tolist()
            logger.error(f"[HSBC] No se encontró columna de fecha. Disponibles: {available}")
            raise ValueError(f"[HSBC] Estructura no reconocida. Esperado header en fila 0.")

        # Filtrar filas con fecha válida
        df = df[df[col_fecha_name].astype(str).apply(looks_like_date)].copy()

        # Renombrar columnas para output
        df_out = pd.DataFrame(index=df.index)
        df_out["raw_date"] = df[col_fecha_name].astype(str)
        df_out["account_id"] = account_id
        df_out["bank"] = self.BANK_NAME
        if col_desc_name and col_desc_name in df.columns:
            df_out["description"] = df[col_desc_name].astype(str)
            df_out["description_detail"] = df[col_desc_name].astype(str)
        else:
            df_out["description"] = ""
            df_out["description_detail"] = ""

        # Crédito (depósito) y Débito (retiro)
        df_out["raw_deposit"] = ""
        df_out["raw_withdrawal"] = ""

        for idx, row in df.iterrows():
            try:
                credito = str(row[col_credito_name]).strip() if col_credito_name and col_credito_name in row.index else ""
                debito = str(row[col_debito_name]).strip() if col_debito_name and col_debito_name in row.index else ""
                
                # Si hay crédito, es depósito
                if credito and credito.lower() not in ("", "nan", "none"):
                    df_out.at[idx, "raw_deposit"] = credito
                # Si hay débito, es retiro (puede ser negativo)
                if debito and debito.lower() not in ("", "nan", "none"):
                    df_out.at[idx, "raw_withdrawal"] = debito
            except Exception:
                pass

        logger.info("[HSBC] Filas parseadas: %d", len(df_out))
        return df_out.reset_index(drop=True)

    @staticmethod
    def _extract_account(raw: pd.DataFrame) -> str:
        # Fila 1 (row 1), columna 2 (col 1)
        try:
            val = str(raw.iloc[1, 1]).strip()
            return val if val.lower() not in ("", "nan", "none") else "UNKNOWN"
        except Exception:
            return "UNKNOWN"


# ════════════════════════════════════════════════════════════
# PARSER PRINCIPAL (auto-detección)
# ════════════════════════════════════════════════════════════

class BankParser:
    """
    Parser unificado. Detecta el formato automáticamente por las columnas
    del CSV y delega al sub-parser correcto.

    Uso:
        df_raw = BankParser().parse("estado_cuenta.csv")
    """

    # Registro de sub-parsers. Para agregar un banco nuevo:
    # 1. Crear clase _MiBancoParser(_BaseBankParser)
    # 2. Añadirla aquí
    _PARSERS: list[type[_BaseBankParser]] = [
        _BBVAParser,
        _BanorteParser,
        _ScotiabankParser,
        _NetPayParser,
        _MercadoPagoParser,
        _HSBCParser,
    ]

    def parse(self, file_path: str) -> pd.DataFrame:
        logger.info("Parseando banco: %s", file_path)

        # Detección especial REPORTE: xlsx con hoja "Hoja2" y columna "TPV"
        path_lower = str(file_path).lower()
        if path_lower.endswith((".xlsx", ".xls")):
            try:
                xl = pd.ExcelFile(file_path)

                # REPORTE CAJA: intentar primero por nombre de archivo,
                # luego por contenido (independientemente de las hojas presentes).
                _fname_upper = str(file_path).upper()
                _force_reporte_caja = (
                    "REPORTE CAJA" in _fname_upper
                    or "REPORTE_CAJA" in _fname_upper
                )

                hoja1 = pd.read_excel(
                    file_path, sheet_name=0,
                    header=None, dtype=str
                ).fillna("")

                _is_reporte_caja = False
                if _force_reporte_caja:
                    _is_reporte_caja = True
                else:
                    # Detectar si alguna de las primeras filas tiene cabeceras de caja
                    _keywords = {"FECHA", "TIENDA", "BANCO", "MONTO"}
                    for _i in range(min(10, len(hoja1))):
                        _row_vals = {str(v).strip().upper() for v in hoja1.iloc[_i].values}
                        if len(_row_vals & _keywords) >= 3:
                            _is_reporte_caja = True
                            break
                    # También detectar por A1 == "TIENDAS" (formato anterior)
                    if not _is_reporte_caja and len(hoja1) > 0:
                        _is_reporte_caja = str(hoja1.iloc[0, 0]).strip().upper() == "TIENDAS"

                if _is_reporte_caja:
                    logger.info("Formato bancario detectado: REPORTE_CAJA (hoja=0)")
                    result = _ReporteCajaParser().parse_raw(hoja1)
                    logger.info("[REPORTE_CAJA] Filas parseadas: %d", len(result))
                    return result

                # REPORTE (banco enriquecido con Hoja2 + columna TPV)
                if "Hoja2" in xl.sheet_names:
                    hoja2 = pd.read_excel(
                        file_path, sheet_name="Hoja2",
                        header=None, dtype=str
                    ).fillna("")
                    if len(hoja2) > 0 and len(hoja2.columns) > 13:
                        if "TPV" in str(hoja2.iloc[0, 13]).upper():
                            logger.info("Formato bancario detectado: REPORTE")
                            result = _ReporteParser().parse_raw(hoja2)
                            logger.info("[REPORTE] Filas parseadas: %d", len(result))
                            return result

                # NETPAY: archivo con hoja 'Ventas Tarjeta Presente'
                if "Ventas Tarjeta Presente" in xl.sheet_names:
                    logger.info("Formato bancario detectado: NETPAY")
                    netpay_raw = pd.read_excel(
                        file_path, sheet_name="Ventas Tarjeta Presente",
                        header=None, dtype=str
                    ).fillna("")
                    result = _NetPayParser().parse_raw(netpay_raw)
                    logger.info("[NETPAY] Filas parseadas: %d", len(result))
                    return result

                # MERCADO PAGO: Lee colores de Excel directamente
                # Detectar por presencia de columna "Total a recibir" o "Número de operación"
                try:
                    _primary_sheet = xl.worksheets[0]
                    _header_row_idx = 3  # Fila 4 (0-indexed)
                    if _header_row_idx < len(_primary_sheet._cells):
                        _hdr_text = " ".join(
                            str(cell.value or "").upper()
                            for cell in list(_primary_sheet.iter_rows(
                                min_row=_header_row_idx + 1,
                                max_row=_header_row_idx + 1,
                                values_only=False
                            ))[:10]
                        )
                        if ("NUMERO DE OPERACION" in _hdr_text or "NÚMERO DE OPERACIÓN" in _hdr_text or
                            "TARJETA PRESENTE" in _hdr_text or "COBRO" in _hdr_text):
                            logger.info("Formato bancario detectado: MERCADO PAGO (por xlsx)")
                            result = _MercadoPagoParser().parse(file_path)
                            logger.info("[MERCADOPAGO] Filas parseadas: %d", len(result))
                            return result
                except Exception as e:
                    logger.debug("[MERCADOPAGO detection] Excepción (no es MP): %s", str(e))
                    pass

            except Exception:
                pass  # no es REPORTE, continuar con detección normal

        raw = _BaseBankParser._read_file(file_path)
        fmt = self._detect_format(raw)

        if fmt is None:
            raise ValueError(
                f"Formato bancario no reconocido en: {file_path}\n"
                "Bancos soportados: BBVA, Banorte/Bancomer."
            )

        logger.info("Formato bancario detectado: %s", fmt.BANK_NAME)
        
        # Mercado Pago requiere acceso a openpyxl para colores, no usa parse_raw()
        if fmt == _MercadoPagoParser:
            result = _MercadoPagoParser().parse(file_path)
        else:
            result = fmt().parse_raw(raw)
        
        logger.info("[%s] Filas parseadas: %d", fmt.BANK_NAME, len(result))
        return result

    @staticmethod
    def _detect_format(raw: pd.DataFrame) -> type[_BaseBankParser] | None:
        if raw.empty:
            return None

        # Leer las 2 primeras filas como texto plano para detección
        row0 = " ".join(raw.iloc[0].fillna("").astype(str).tolist())
        row1 = " ".join(raw.iloc[1].fillna("").astype(str).tolist()) if len(raw) > 1 else ""

        row0_upper = _normalize_text(row0)
        row1_upper = _normalize_text(row1)

        # BBVA: header en fila 0 con DEPÓSITOS / RETIROS
        if "DEPOSITOS" in row0_upper or "FECHA DE OPERACI" in row0_upper:
            return _BBVAParser

        # BANORTE: fila 0 comienza con "Cuenta" y fila 1 tiene "Cargo" / "Abono"
        if str(raw.iloc[0, 0]).strip().lower() == "cuenta" and (
            "CARGO" in row1_upper or "ABONO" in row1_upper
        ):
            return _BanorteParser

        # SCOTIABANK: datos directamente desde fila 0, col 7 tiene "Cargo" o "Abono"
        if len(raw.columns) > 7 and str(raw.iloc[0, 7]).strip() in ("Cargo", "Abono"):
            return _ScotiabankParser

        # NETPAY: busca en las primeras 30 filas alguna que contenga
        # "Monto de Trx" / "Monto TRX" o "Fecha de Movimiento" + "Monto"
        _np_text = " ".join(
            raw.iloc[:min(30, len(raw))].fillna("").astype(str).values.flatten()
        ).upper()
        if "MONTO DE TRX" in _np_text or "MONTO TRX" in _np_text or "FECHA DE MOVIMIENTO" in _np_text:
            return _NetPayParser

        # MERCADO PAGO: fila 0 contiene "Ventas" o fila 3 col 0 = "Número de operación"
        row0_text = " ".join(raw.iloc[0].fillna("").astype(str).tolist()).upper()
        if "MERCADO PAGO" in row0_text or "VENTAS" in row0_text:
            return _MercadoPagoParser
        if len(raw) > 3 and "operaci" in str(raw.iloc[3, 0]).lower():
            return _MercadoPagoParser

        # HSBC: encabezado con número de cuenta + pistas de layout HSBC.
        if ("NUMERO DE CUENTA" in row0_upper) and (
            "HSBC" in row0_upper or "FECHA VALOR" in row0_upper or "IMPORTE DE CREDITO" in row0_upper
            or "IMPORTE DEL DEBITO" in row0_upper
        ):
            return _HSBCParser

        return None
