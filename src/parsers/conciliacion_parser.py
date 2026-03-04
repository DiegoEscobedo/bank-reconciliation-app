"""
conciliacion_parser.py — Lee el archivo 'Conciliación Bancaria' mensual Excel
y extrae los ítems pendientes (Más/Menos) de cada hoja de cuenta.

Estructura esperada del Excel (por hoja de cuenta):
  ─ Filas 1-15  : encabezado
  ─ Fila 16     : "Más:" en col I (idx 8) — inicio sección JDE
  ─ Filas 17-N  : pendientes  col I=descripción, col M=monto, col J(ó K)=tipo
  ─ Fila N+2    : "Menos:" en col I — inicio sección banco
  ─ Filas N+3-M : pendientes  mismo formato
  ─ Fila M+2    : "Otros:" — fin de pendientes

Formatos de fecha en col I:
  'DD-mmm-YY'   → '29-feb-24', '11-mar-25', '05-en-26'
  'DD-mm-YY'    → '08-01-26', '12-01-26'
  'DD-mmm-YYYY' → '18-feb-2025'
  Typos habituales: '29-0ct-25' (cero en lugar de 'o'), '31-oct25' (guion faltante)
"""

import re
import logging
import datetime as _dt
from typing import Union

import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════
# MAPEO DE MESES EN ESPAÑOL
# ═══════════════════════════════════════════════════════════

_MESES_ES: dict[str, int] = {
    "en": 1,  "ene": 1,  "enero": 1,
    "feb": 2, "febrero": 2,
    "mar": 3, "marzo": 3,
    "abr": 4, "abril": 4,
    "may": 5, "mayo": 5,
    "jun": 6, "junio": 6,
    "jul": 7, "julio": 7,
    "ago": 8, "agto": 8, "agosto": 8,   # 'agto' es abreviatura usada frecuentemente
    "sep": 9, "sept": 9, "septiembre": 9,
    "oct": 10, "0ct": 10,  # typo frecuente: cero en lugar de 'o'
    "noviembre": 11, "nov": 11,
    "dic": 12, "diciembre": 12,
}

# Regex: DD-mesTexto-YY[YY]  (acepta - / y . como separador)
_RE_MES_TEXTO = re.compile(
    r"^(\d{1,2})[-/.]([a-zA-Z0]{2,})[-/.\s]?(\d{2,4})",
    re.IGNORECASE,
)
# Regex: DD-MM-YY[YY]
_RE_NUMERICO = re.compile(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})")
# Regex: para el caso donde falta el guion: '31oct25'
_RE_SIN_GUION = re.compile(r"^(\d{1,2})([a-zA-Z]{2,})(\d{2,4})", re.IGNORECASE)
# Regex: ISO datetime string 'YYYY-MM-DD ...' (celda de tipo fecha en Excel)
_RE_ISO = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")

# Marcadores de sección (minúsculas)
_MAS_MARKERS    = {"más:", "mas:", "más", "mas"}
_MENOS_MARKERS  = {"menos:", "menos"}
_OTROS_MARKERS  = {"otros:", "otros"}

# Hojas a ignorar
_SKIP_SHEETS = {"SALDOS", "RESUMEN"}


# ═══════════════════════════════════════════════════════════
# PARSEO DE FECHA
# ═══════════════════════════════════════════════════════════

def _parse_concil_date(text: str) -> pd.Timestamp:
    """
    Extrae la fecha del texto de descripción de un pendiente.

    El texto tiene la forma  'DD-mmm-YY TIENDA TIPO'  ó  'DD-MM-YY TIENDA'.
    Retorna pd.NaT si no puede parsear.
    """
    if not text:
        return pd.NaT
    s = text.strip()

    # 1) Mes textual: DD-mmm-YY[YY]
    m = _RE_MES_TEXTO.match(s)
    if m:
        day      = int(m.group(1))
        mes_str  = m.group(2).lower()
        year_str = m.group(3)
        month    = _MESES_ES.get(mes_str)
        if month:
            year = int(year_str)
            if year < 100:
                year += 2000
            try:
                return pd.Timestamp(year, month, day)
            except Exception:
                pass

    # 2) Numérico: DD-MM-YY[YY]
    m2 = _RE_NUMERICO.match(s)
    if m2:
        day   = int(m2.group(1))
        month = int(m2.group(2))
        year  = int(m2.group(3))
        if year < 100:
            year += 2000
        try:
            return pd.Timestamp(year, month, day)
        except Exception:
            pass

    # 3) Sin guion: '31oct25'
    m3 = _RE_SIN_GUION.match(s)
    if m3:
        day      = int(m3.group(1))
        mes_str  = m3.group(2).lower()
        year_str = m3.group(3)
        month    = _MESES_ES.get(mes_str)
        if month:
            year = int(year_str)
            if year < 100:
                year += 2000
            try:
                return pd.Timestamp(year, month, day)
            except Exception:
                pass

    # 4) ISO datetime string: '2026-01-09 00:00:00' (valor de celda de tipo fecha en Excel)
    m4 = _RE_ISO.match(s)
    if m4:
        try:
            return pd.Timestamp(int(m4.group(1)), int(m4.group(2)), int(m4.group(3)))
        except Exception:
            pass

    return pd.NaT


# ═══════════════════════════════════════════════════════════
# PARSER PRINCIPAL
# ═══════════════════════════════════════════════════════════

def parse_conciliacion_excel(file_path: Union[str, bytes]) -> pd.DataFrame:
    """
    Lee el archivo de 'Conciliación Bancaria' mensual y extrae todos
    los ítems pendientes de cada hoja de cuenta.

    Parameters
    ----------
    file_path : str | bytes
        Ruta al archivo .xlsx o bytes del archivo (para Streamlit upload).

    Returns
    -------
    pd.DataFrame con columnas:
        account_id      str   – nombre de la hoja (número de cuenta, ej: '6614')
        section         str   – 'mas'  = en JDE no en banco
                                'menos' = en banco no en JDE
        movement_date   Timestamp | NaT
        description     str   – texto completo de la celda (incluyendo prefijo de fecha)
        abs_amount      float – monto positivo
        type_code       int | None – código de tipo de movimiento (col J ó K)
    """
    # --- Cargar workbook ---------------------------------------------------
    try:
        if isinstance(file_path, (bytes, bytearray)):
            import io
            wb = openpyxl.load_workbook(
                io.BytesIO(file_path), data_only=True, read_only=True
            )
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(
            f"No se pudo abrir el archivo de conciliación: {exc}"
        ) from exc

    all_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().upper() in _SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        section: str | None = None
        mas_seen   = False
        menos_seen = False

        for row in ws.iter_rows():
            # Recolectar celdas con valor (0-indexed col)
            cells: dict[int, object] = {
                cell.column - 1: cell.value
                for cell in row
                if cell.value is not None
            }
            if not cells:
                continue

            # ── Columna I (idx 8): descripción / marcador ─────────────────
            raw_i = cells.get(8)
            # Si Excel devuelve un objeto datetime/date directamente, convertir a string
            # para que el parser de fecha lo detecte con el regex ISO
            if isinstance(raw_i, (_dt.datetime, _dt.date)):
                raw_i = str(pd.Timestamp(raw_i))
            desc    = str(raw_i).strip() if raw_i is not None else ""
            desc_lo = desc.lower()

            # Detectar marcador "Más:"
            if desc_lo in _MAS_MARKERS:
                section    = "mas"
                mas_seen   = True
                continue

            # Detectar marcador "Menos:"
            if desc_lo in _MENOS_MARKERS and mas_seen:
                section    = "menos"
                menos_seen = True
                continue

            # Fin de sección pendientes
            if desc_lo in _OTROS_MARKERS:
                break
            if "saldo" in desc_lo and section is not None:
                break

            # Sin sección activa → ignorar
            if section is None:
                continue

            # ── Columna M (idx 12): monto ─────────────────────────────────
            raw_m = cells.get(12)

            # Saltar si es fórmula (str con '=') o None
            if raw_m is None:
                continue
            if isinstance(raw_m, str):
                if "=" in raw_m:
                    continue  # fórmula SUM, es el subtotal
                try:
                    raw_m = float(raw_m)
                except ValueError:
                    continue

            try:
                abs_amount = abs(float(raw_m))
            except (TypeError, ValueError):
                continue

            if abs_amount <= 0:
                continue

            # Saltar si descripción está vacía o es un marcador
            if not desc or desc_lo in _MAS_MARKERS | _MENOS_MARKERS | _OTROS_MARKERS:
                continue

            # ── Código de tipo (col J idx 9 ó K idx 10) ──────────────────
            raw_type = cells.get(9) or cells.get(10)
            try:
                type_code: int | None = int(float(str(raw_type))) if raw_type is not None else None
            except (TypeError, ValueError):
                type_code = None

            # ── Parsear fecha ─────────────────────────────────────────────
            movement_date = _parse_concil_date(desc)

            all_rows.append({
                "account_id":    sheet_name,
                "section":       section,
                "movement_date": movement_date,
                "description":   desc,
                "abs_amount":    abs_amount,
                "type_code":     type_code,
            })

    wb.close()

    if not all_rows:
        return pd.DataFrame(
            columns=[
                "account_id", "section", "movement_date",
                "description", "abs_amount", "type_code",
            ]
        )

    df = pd.DataFrame(all_rows)
    df["movement_date"] = pd.to_datetime(df["movement_date"], errors="coerce")
    df["abs_amount"]    = pd.to_numeric(df["abs_amount"], errors="coerce")
    return df


# ═══════════════════════════════════════════════════════════
# RESUMEN / ESTADÍSTICAS
# ═══════════════════════════════════════════════════════════

def get_pending_summary(df: pd.DataFrame) -> dict:
    """
    Retorna un dict con estadísticas del DataFrame de pendientes históricos.
    """
    if df.empty:
        return {
            "total": 0,
            "mas_count": 0,
            "menos_count": 0,
            "mas_total": 0.0,
            "menos_total": 0.0,
            "accounts": [],
            "date_min": None,
            "date_max": None,
        }

    mas   = df[df["section"] == "mas"]
    menos = df[df["section"] == "menos"]

    return {
        "total":       len(df),
        "mas_count":   len(mas),
        "menos_count": len(menos),
        "mas_total":   float(mas["abs_amount"].sum()),
        "menos_total": float(menos["abs_amount"].sum()),
        "accounts":    sorted(df["account_id"].unique().tolist()),
        "date_min":    df["movement_date"].min(),
        "date_max":    df["movement_date"].max(),
    }
