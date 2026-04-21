import io
import unittest

import openpyxl
import pandas as pd

from src.reporting.excel_reporter import ExcelReporter


class TestWriteBackStrictSign(unittest.TestCase):
    @staticmethod
    def _build_pt_bytes() -> bytes:
        buffer = io.BytesIO()

        # xlsxwriter genera sharedStrings, requeridas por el parser XML del write-back.
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            df = pd.DataFrame(
                [
                    {
                        "Aux_Fact": "A1",
                        "Descripción": "CUENTA 6614 MOV PRUEBA",
                        "Importe": "-100.00",
                        "CONCILIADO": "",
                        "FECHA CONCILIACION": "",
                    }
                ]
            )
            df.to_excel(writer, index=False, sheet_name="AUX CONTABLE")

        return buffer.getvalue()

    def test_strict_writeback_rejects_amount_sign_mismatch(self):
        reporter = ExcelReporter()
        source_bytes = self._build_pt_bytes()

        with self.assertRaises(ValueError):
            reporter.write_back_conciliados(
                source=source_bytes,
                reconciled_aux_facts=["A1"],
                filter_accounts=["6614"],
                strict_match_entries=[
                    {
                        "aux_fact": "A1",
                        "account_id": "6614",
                        # Signo contrario al del Papel de Trabajo.
                        "amount_signed": 100.0,
                    }
                ],
                amount_tolerance=0.5,
                require_full_strict_match=True,
            )

    def test_strict_writeback_accepts_same_amount_sign(self):
        reporter = ExcelReporter()
        source_bytes = self._build_pt_bytes()

        out_bytes = reporter.write_back_conciliados(
            source=source_bytes,
            reconciled_aux_facts=["A1"],
            filter_accounts=["6614"],
            strict_match_entries=[
                {
                    "aux_fact": "A1",
                    "account_id": "6614",
                    "amount_signed": -100.0,
                }
            ],
            amount_tolerance=0.5,
            require_full_strict_match=True,
        )

        wb = openpyxl.load_workbook(io.BytesIO(out_bytes), data_only=True)
        ws = wb["AUX CONTABLE"]

        headers = [cell.value for cell in ws[1]]
        conc_col = headers.index("CONCILIADO") + 1
        fecha_col = headers.index("FECHA CONCILIACION") + 1

        self.assertEqual(ws.cell(row=2, column=conc_col).value, "SÍ")
        self.assertTrue(ws.cell(row=2, column=fecha_col).value)


if __name__ == "__main__":
    unittest.main()
